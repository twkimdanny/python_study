from openpyxl import Workbook
from openpyxl import load_workbook

# 엑셀파일 쓰기
write_wb = Workbook()

write_ws = write_wb.create_sheet('테스트1')
# write_ws = write_wb.active
write_ws = write_wb['테스트1']
write_ws['A1'] = '학번'
write_ws['B1'] = '이름'
write_ws['A2'] = '202311111'
write_ws['B2'] = '홍길동'
write_wb.save('aaa.xlsx')

# 엑셀파일 읽기
load_wb = load_workbook('aaa.xlsx', data_only=True)
# 테스트1 시트 가져오기
load_ws = load_wb['테스트1']
# print(load_ws['A2'])
# hakbun = str(load_ws['A2'].value)
# print(load_ws['A2'].value)
get_cells = load_ws['A2':'B2']
for row in get_cells :
    for cell in row :
        print(cell.value)